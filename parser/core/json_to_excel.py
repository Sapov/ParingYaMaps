import json
from openpyxl import Workbook
from openpyxl.styles import Font


def create_excel_file(name_file: str, lst_item: list) -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = name_file

    # Заголовки
    headers = [
        'Ссылка', 'Название', 'Рейтинг', 'Оценки', 'Адрес',
        'Телефон', 'Сайт', 'Email', 'Whatsapp', 'Telega'
    ]

    for col, header in enumerate(headers, start=1):
        ws.cell(1, col, header)

    # Жирный шрифт для заголовков
    ft = Font(bold=True)
    for row in ws["A1:J1"]:
        for cell in row:
            cell.font = ft

    row_num = 2  # Начинаем с 2 строки (после заголовков)

    for item in lst_item:
        if not item:  # Пропускаем None или пустые элементы
            continue

        # Безопасное получение значений с defaults
        link = item.get('link', '')
        name = item.get('name', '')
        rating = item.get('rating_yandex', '')
        estimation = item.get('estimation', '')
        address = item.get('address', '')
        phone = item.get('phone', '')
        site = item.get('site', '')

        # Заполняем ячейки
        ws[f'A{row_num}'] = link
        ws[f'B{row_num}'] = name
        ws[f'C{row_num}'] = rating
        ws[f'D{row_num}'] = estimation
        ws[f'E{row_num}'] = address
        ws[f'F{row_num}'] = phone
        ws[f'G{row_num}'] = site

        # Обработка mail (может отсутствовать или быть не списком)
        mail = item.get('mail')
        if isinstance(mail, list):
            # Фильтруем пустые строки в списке mail
            mail_filtered = [m for m in mail if m]
            ws[f'H{row_num}'] = ', '.join(str(m) for m in mail_filtered) if mail_filtered else ''
        elif mail:  # если это строка или число
            ws[f'H{row_num}'] = str(mail)

        # Обработка whatsapp
        whatsapp = item.get('whatsapp')
        if isinstance(whatsapp, list):
            whatsapp_filtered = [w for w in whatsapp if w]
            ws[f'I{row_num}'] = ', '.join(str(w) for w in whatsapp_filtered) if whatsapp_filtered else ''
        elif whatsapp:
            ws[f'I{row_num}'] = str(whatsapp)

        # Обработка telegram
        telegram = item.get('telegram')
        if isinstance(telegram, list):
            telegram_filtered = [t for t in telegram if t]
            ws[f'J{row_num}'] = ', '.join(str(t) for t in telegram_filtered) if telegram_filtered else ''
        elif telegram:
            ws[f'J{row_num}'] = str(telegram)

        row_num += 1  # Увеличиваем номер строки только для непустых элементов

    wb.save(name_file)



def open_json(json_file: str) -> list[dict]:
    with open(json_file, 'r', encoding='utf-8') as file:
        return json.load(file)


def main(name_file: str):
    lst = open_json('scv_json/temp_mail.json')
    create_excel_file(f'{name_file}.xlsx', lst)  # Исправлено на .xlsx


if __name__ == '__main__':
    main('Оптовые базы')